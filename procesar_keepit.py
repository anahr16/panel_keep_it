#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
procesar_keepit.py — Actualiza el panel DATA KEEP IT 2024.xlsx hasta la fecha.

Toma los insumos de Consolidadores/ y agrega a las 3 hojas operativas
(Casos Keep it, Cupones Keep it, Ahorro) las filas de los meses nuevos, dejando
calculado el AHORRO con la receta validada. Genera un archivo NUEVO; NO toca el original.

Receta de ahorro (validada 94%):
    Precio Costo      = Precio Unitario ($) del archivo de costos (por SKU F.COM y mes)
    Perdida costo     = Precio Costo * FACTOR_PERDIDA (0.5)
    Costo cupon       = Monto del cupon * FACTOR_COSTO_CUPON (0.7)
    Logistica inversa = tarifa mensual fija (2026 = 15.847)
    Ahorro            = Perdida costo - Costo cupon + Logistica inversa
"""
import re
from pathlib import Path
import pandas as pd
import numpy as np

BASE = Path(__file__).parent
CONS = BASE / "Consolidadores"

# --- Parametros de negocio (CONFIRMAR con el negocio) ---
FACTOR_PERDIDA      = 0.50     # % del costo que se pierde (recuperacion 50%)
FACTOR_COSTO_CUPON  = 0.70     # costo real del cupon = 70% del monto
LOGISTICA_INVERSA   = 15847    # tarifa mensual fija vigente 2026 (CONFIRMAR para abr-jun)

# 6 motivos que SIEMPRE entran (auto-marcan el flag Keep It)
TIPIFICACIONES_KEEPIT = {
    "Producto en mal estado", "Producto incompleto", "Pieza faltante",
    "Producto con falla de funcionamiento/técnica",
    "Producto Pieza Rota", "Producto con empaque deteriorado",
}
# Estas tipificaciones SOLO entran si el caso tiene el flag Keep It == 1
TIPIFICACIONES_SOLO_FLAG = {"Producto no corresponde"}

ARCH = {
    "panel":   CONS / "DATA KEEP IT 2024.xlsx",
    "bot":     CONS / "Cupones BOT - Keep It 2023.xlsx",
    "jerarq":  CONS / "2279dfef-30f3-4dee-9346-d31aea31ad2f.xlsx",
    "costo":   CONS / "ad4a57bf-3881-46a3-9e6d-72a914f77766.xlsx",
    "casos_may": CONS / "Informe Productos-Reserva Keep-it-2026-06-23-09-41-29.xlsx",
    "casos_jun": CONS / "Informe Productos-Reserva Keep-it-2026-06-23-11-47-34.xlsx",
}

def limpia_prefijo(x):
    if pd.isna(x): return x
    return re.sub(r"^\d+\s*-\s*", "", str(x)).strip()

def norm_sku(x):
    """Normaliza SKU: quita '.0' de floats y espacios."""
    if pd.isna(x): return None
    s = str(x).strip()
    return s[:-2] if s.endswith(".0") else s

def monto_desde_codigo(codigo, idpromo=None):
    """Extrae el monto del codigo: 'KEEP10M..'->10000, '_55M-'->55000, 'KEEP25.29'->25000."""
    for txt in (codigo, idpromo):
        if pd.isna(txt): continue
        m = re.search(r"(\d{1,3})\s*[M.]", str(txt))   # denominacion seguida de 'M' o '.'
        if m: return int(m.group(1)) * 1000
    return np.nan

# ====================================================================
# 1) Maestros de lookup
# ====================================================================
print("Cargando maestros...")
jer = pd.read_excel(ARCH["jerarq"])
jer["_fcom"] = jer["SKU F.COM"].astype(str).str.strip()
jer["_sku"]  = jer["SKU"].astype(str).str.strip()
jer["Familia_l"]    = jer["Familia"].map(limpia_prefijo)
jer["SubFamilia_l"] = jer["Sub Familia"].map(limpia_prefijo)
fam_by_fcom = jer.drop_duplicates("_fcom").set_index("_fcom")[["Familia_l","SubFamilia_l"]]
fcom_a_sku  = jer.drop_duplicates("_fcom").set_index("_fcom")["_sku"]

cost = pd.read_excel(ARCH["costo"])
cost["_fcom"] = cost["SKU F.COM"].astype(str).str.strip()
cost["_ym"]   = pd.to_datetime(cost["Año Mes"], errors="coerce").dt.to_period("M").astype(str)
cost["_precio"] = pd.to_numeric(cost["Precio Unitario ($)"], errors="coerce")
cost_key = cost.dropna(subset=["_precio"]).drop_duplicates(["_fcom","_ym"]).set_index(["_fcom","_ym"])["_precio"]
cost_any = cost.dropna(subset=["_precio"]).drop_duplicates("_fcom").set_index("_fcom")["_precio"]

def precio_costo(fcom, ym):
    v = cost_key.get((fcom, ym), np.nan)
    if pd.isna(v): v = cost_any.get(fcom, np.nan)
    return v

cas0 = pd.read_excel(ARCH["panel"], "Casos Keep it")
cup0 = pd.read_excel(ARCH["panel"], "Cupones Keep it")
aho0 = pd.read_excel(ARCH["panel"], "Ahorro")
sku0 = pd.read_excel(ARCH["panel"], "SKU")

port_map = {}
for df in (cas0, cup0):
    for k,v in zip(df["SKU"].astype(str).str.strip(), df["Portable"]):
        if pd.notna(v) and v not in ("#N/D","nan") and k not in port_map:
            port_map[k]=v
name_map = {}
for k,v in zip(cas0["SKU"].astype(str).str.strip(), cas0["Nombre del producto"]):
    if pd.notna(v) and k not in name_map: name_map[k]=v

MESES_ES = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",7:"Julio",
            8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}

# ====================================================================
# 2) CASOS
# ====================================================================
print("Procesando CASOS...")
exp_full = pd.concat([pd.read_excel(ARCH["casos_may"], header=14),
                      pd.read_excel(ARCH["casos_jun"], header=14)], ignore_index=True).dropna(how="all")
exp_full = exp_full[[c for c in exp_full.columns if not str(c).startswith("Unnamed")]]

# Mapas Caso -> SKU (F.COM) y Caso -> Tipificación; el BOT 2026 no trae estos campos
caso_sku, caso_tip = {}, {}
for df in (exp_full, cas0):
    for k, v, t in zip(pd.to_numeric(df["Número del caso"], errors="coerce"),
                       df["SKU"], df["Nombre Tipificación"]):
        ik = int(k) if pd.notna(k) else None
        if ik is None: continue
        sv = norm_sku(v)
        if sv and ik not in caso_sku: caso_sku[ik] = sv
        if pd.notna(t) and ik not in caso_tip: caso_tip[ik] = t

_flag = pd.to_numeric(exp_full["Caso Keep It"], errors="coerce")
_tip  = exp_full["Nombre Tipificación"]
exp = exp_full[_tip.isin(TIPIFICACIONES_KEEPIT) |
               (_tip.isin(TIPIFICACIONES_SOLO_FLAG) & (_flag == 1))]
exp["Número del caso"] = pd.to_numeric(exp["Número del caso"], errors="coerce")
exp = exp.dropna(subset=["Número del caso"]).drop_duplicates("Número del caso")
ya = set(pd.to_numeric(cas0["Número del caso"], errors="coerce").dropna().astype("int64"))
exp = exp[~exp["Número del caso"].astype("int64").isin(ya)]

casos_cols = list(cas0.columns)
nc = pd.DataFrame(index=exp.index, columns=casos_cols)
for col in casos_cols[:32]:
    if col in exp.columns: nc[col] = exp[col].values
# las fechas del export vienen como texto d/m/a -> a datetime real (como el panel viejo)
for dcol in ["Fecha/Hora de apertura", "Fecha/Hora de cierre", "Fecha de compra"]:
    if dcol in nc.columns:
        nc[dcol] = pd.to_datetime(nc[dcol], errors="coerce", dayfirst=True)
sk = exp["SKU"].astype(str).str.strip()
ap = pd.to_datetime(exp["Fecha/Hora de apertura"], errors="coerce", dayfirst=True)
nc["Portable"]            = sk.map(port_map).fillna("#N/D").values
nc["Familia"]             = sk.map(fam_by_fcom["Familia_l"]).fillna("#N/D").values
nc["Subfamilia"]          = sk.map(fam_by_fcom["SubFamilia_l"]).fillna("#N/D").values
nc["Mes"]                 = ap.dt.month.values
nc["Nombre mes"]          = [f"{MESES_ES.get(m,'')} - {y}" if pd.notna(m) else "" for m,y in zip(ap.dt.month, ap.dt.year)]
nc["Descripcion producto"]= exp["Nombre del producto"].values
nc["Número de reserva 2"] = exp["Número de reserva"].values
nc["Tipo de reserva 2"]   = exp["Tipo de reserva"].values
casos_final = pd.concat([cas0, nc], ignore_index=True)

# ====================================================================
# 3) CUPONES
# ====================================================================
print("Procesando CUPONES...")
bot = pd.read_excel(ARCH["bot"], "2026")
bot.columns = [str(c).strip() for c in bot.columns]
bot["Fecha de entrega"] = pd.to_datetime(bot["Fecha de entrega"], errors="coerce")
ya_cod = set(cup0["Codigo"].astype(str).str.strip())
bot = bot[(~bot["Codigo"].astype(str).str.strip().isin(ya_cod)) & bot["Fecha de entrega"].notna()]

cup_cols = list(cup0.columns)
mapeo_bot = {c:c for c in ["Codigo","Monto_Bot","Monto","Compra Min","Fecha Ini","Fecha Fin",
    "ID Promo","Descripcion","Propietario","Caso SF","Ejecutivo","Equipo","Fecha de entrega",
    "Rut Cliente","Monto compra","Motivo","Solucion","Producto","Reserva"]}
ncp = pd.DataFrame(index=bot.index, columns=cup_cols)
for p,b in mapeo_bot.items():
    if b in bot.columns: ncp[p] = bot[b].values
fe  = bot["Fecha de entrega"]
caso = pd.to_numeric(bot["Caso SF"], errors="coerce")
# SKU desde el caso (el BOT 2026 no lo trae); Monto desde el codigo si falta
skb = caso.map(lambda k: caso_sku.get(int(k)) if pd.notna(k) else None)
monto = pd.to_numeric(bot["Monto"], errors="coerce")
monto = monto.where(monto.notna(),
                    [monto_desde_codigo(c, i) for c, i in zip(bot["Codigo"], bot.get("ID Promo"))])
ncp["Monto"]         = monto.values
# el BOT 2026 no trae Compra Min -> regla del histórico: Compra Min = Monto + 5.000
ncp["Compra Min"]    = (monto + 5000).values
ncp["SKU"]           = skb.values
ncp["Año"]           = fe.dt.year.values
ncp["mes2"]          = fe.dt.month.values
ncp["Mes"]           = fe.dt.month.values
ncp["Familia_1"]     = skb.map(fam_by_fcom["Familia_l"]).fillna("#N/D").values
ncp["Subfamilia"]    = skb.map(fam_by_fcom["SubFamilia_l"]).fillna("#N/D").values
ncp["Portable"]      = skb.map(port_map).fillna("#N/D").values
ncp["Nombnre producto"] = skb.map(name_map).values
cupones_final = pd.concat([cup0, ncp], ignore_index=True)

# ====================================================================
# 4) AHORRO
# ====================================================================
print("Procesando AHORRO...")
aho_cols = list(aho0.columns)
nb = bot.copy()
nb["_caso"] = pd.to_numeric(nb["Caso SF"], errors="coerce")
nb["_fcom"] = nb["_caso"].map(lambda k: caso_sku.get(int(k)) if pd.notna(k) else None)
nb["_ym"]   = nb["Fecha de entrega"].dt.to_period("M").astype(str)
nb["Precio Costo"] = [precio_costo(f, y) if f else np.nan for f,y in zip(nb["_fcom"], nb["_ym"])]
nb["Precio Costo"] = nb["Precio Costo"].round(0)          # enteros, como el formato original
nb["Monto_n"]      = monto.values
nb["Perdida costo"]= (nb["Precio Costo"] * FACTOR_PERDIDA).round(0)
nb["Costo cupón"]  = (nb["Monto_n"] * FACTOR_COSTO_CUPON).round(0)
nb["Logisitica inversa"] = LOGISTICA_INVERSA
nb["Ahorro"]       = nb["Perdida costo"] - nb["Costo cupón"] + nb["Logisitica inversa"]

n0 = pd.to_numeric(aho0["Num Cupon"], errors="coerce").max()
nb = nb.reset_index(drop=True)
nha = pd.DataFrame(index=nb.index, columns=aho_cols)
nha["Num Cupon"]        = (int(n0) + 1 + nb.index).values
nha["Fecha de entrega"] = nb["Fecha de entrega"].values
nha["Monto"]            = nb["Monto_n"].values
nha["Rut Cliente"]      = nb["Rut Cliente"].values
nha["Motivo"]           = nb["Motivo"].values
nha["Caso"]             = pd.to_numeric(nb["Caso SF"], errors="coerce").values
nha["SKU"]              = nb["_fcom"].map(fcom_a_sku).fillna(nb["_fcom"]).values
nha["Solicitado por"]   = nb["Ejecutivo"].values
nha["Precio Costo"]     = nb["Precio Costo"].values
nha["Perdida costo"]    = nb["Perdida costo"].values
nha["Logisitica inversa"]= nb["Logisitica inversa"].values
nha["Costo cupón"]      = nb["Costo cupón"].values
nha["Categoria producto"]= nb["_fcom"].map(fam_by_fcom["Familia_l"]).values
nha["Sub familia producto"]= nb["_fcom"].map(fam_by_fcom["SubFamilia_l"]).values
# columnas de clasificación que el chico llenaba (derivables)
nha["Tipo Reclamo"]     = nb["_caso"].map(lambda k: caso_tip.get(int(k)) if pd.notna(k) else None).values
nha["Categoria"]        = "Liquidación"          # constante en todo el histórico
nha["NCR"]              = "cupón"                # etiqueta constante
nha["Margen"]           = 0.30                   # valor por defecto (64% del histórico)
nha["tipo producto"]    = "Otras"                # valor dominante
nha["Ahorro"]           = nb["Ahorro"].values
nha["Año"]              = nb["Fecha de entrega"].dt.year.values
nha["Mes"]              = nb["Fecha de entrega"].dt.month.values
# donde falló el lookup, poner "#N/D" (igual que el original), no vacío
COLS_ND = ["SKU", "Costo Prod.", "Precio Costo", "Perdida costo", "Tipo Reclamo",
           "Categoria producto", "Sub familia producto", "Ahorro"]
for c in COLS_ND:
    nha[c] = nha[c].where(nha[c].notna(), "#N/D")
ahorro_final = pd.concat([aho0, nha], ignore_index=True)

# ====================================================================
# 5) Guardar archivo NUEVO
# ====================================================================
from openpyxl.styles import PatternFill
AZUL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # celdas a agregar

out = CONS / "DATA KEEP IT 2024_actualizado_2026-06-23_v3.xlsx"
print(f"\nEscribiendo {out.name} ...")
with pd.ExcelWriter(out, engine="openpyxl") as w:
    cupones_final.to_excel(w, sheet_name="Cupones Keep it", index=False)
    casos_final.to_excel(w, sheet_name="Casos Keep it", index=False)
    sku0.to_excel(w, sheet_name="SKU", index=False)
    ahorro_final.to_excel(w, sheet_name="Ahorro", index=False)
    # pintar de AZUL las filas nuevas (las que el usuario debe agregar al panel)
    def pintar_nuevas(sheet, n_viejo, n_total, n_cols):
        ws = w.book[sheet]
        for r in range(n_viejo + 2, n_total + 2):   # +1 cabecera, +1 base-1
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = AZUL
    pintar_nuevas("Cupones Keep it", len(cup0), len(cupones_final), cupones_final.shape[1])
    pintar_nuevas("Casos Keep it",   len(cas0), len(casos_final),   casos_final.shape[1])
    pintar_nuevas("Ahorro",          len(aho0), len(ahorro_final),  ahorro_final.shape[1])

# ====================================================================
# 6) Reporte
# ====================================================================
print("\n" + "="*60 + "\nRESUMEN DE LA ACTUALIZACION")
print(f"  Casos:   {len(cas0):>6} -> {len(casos_final):>6}  (+{len(nc)})")
print(f"  Cupones: {len(cup0):>6} -> {len(cupones_final):>6}  (+{len(ncp)})")
print(f"  Ahorro:  {len(aho0):>6} -> {len(ahorro_final):>6}  (+{len(nha)})")
print(f"\n  Cupones nuevos con costo (=> ahorro): {nha['Precio Costo'].notna().sum()}/{len(nha)} ({nha['Precio Costo'].notna().mean()*100:.0f}%)")
print(f"  Ahorro nuevo total: ${pd.to_numeric(nha['Ahorro'],errors='coerce').sum():,.0f}")
print(f"  Familia asignada en casos nuevos: {(nc['Familia']!='#N/D').mean()*100:.0f}%")
print(f"  Meses nuevos cupones: {sorted(fe.dt.to_period('M').astype(str).unique())}")
print(f"  Meses nuevos casos:   {sorted(ap.dt.to_period('M').dropna().astype(str).unique())}")
